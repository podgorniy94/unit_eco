import json
from math import ceil
from tkinter import Event, filedialog, messagebox, simpledialog, ttk
from typing import Final, Optional

import customtkinter as cttk
import openpyxl
import requests
from pyperclip import copy


class Validation:
    mandatory: Final = "Обязательное поле"
    not_digit: Final = "Нечисловое значение"
    digits_seps: Final = "0123456789,."

    @classmethod
    def validate(cls, win, entries: dict) -> bool:

        win.focus_set()
        valid: list = []

        for entry in entries.values():
            mandatory = cls.mandatory
            not_digit = cls.not_digit
            digits_seps = cls.digits_seps

            value = entry.get()

            if "," in value:
                updated_value = value.replace(",", ".")
                entry.delete(0, cttk.END)
                entry.insert(0, updated_value)

            if value == mandatory or value == not_digit:
                continue
            elif not value:
                entry.insert(0, mandatory)
                entry.configure(text_color="red")
            elif not all(digit in digits_seps for digit in value):
                entry.delete(0, cttk.END)
                entry.insert(0, not_digit)
                entry.configure(text_color="red")
            elif value == "0":
                entry.delete(0, cttk.END)
                entry.insert(0, "0")
                entry.configure(text_color="red")
            else:
                valid.append(True)

        return len(valid) == len(entries)

    @classmethod
    def clean(cls, field: cttk.CTkEntry, event: Event) -> None:
        if type(event) == Event:
            value = field.get()
            if value == cls.mandatory or value == cls.not_digit or value == "0":
                field.configure(text_color="black")
                field.delete(0, cttk.END)


class Calculation(Validation):
    @classmethod
    def calculate(cls, win, entries: dict, buttons: dict):
        if cls.validate(win, entries):
            density_args = (
                entries["product_amount"],
                entries["amount_in_box"],
                entries["box_weight"],
                entries["height"],
                entries["width"],
                entries["length"],
            )
            density_args = [float(i.get()) for i in density_args]
            costs_args = (
                entries["price"],
                entries["product_amount"],
                entries["currency"],
            )
            price, product_amount, currency = [float(i.get()) for i in costs_args]

            bts = (
                buttons["goods_cost"],
                buttons["package"],
                buttons["insurance"],
                buttons["total_cost"],
                buttons["one_goods"],
                buttons["cargo"],
            )

            density, total_weight, total_volume = cls.calculate_density(*density_args)
            cargo = cls.calculate_cargo(density, total_weight, total_volume, currency)
            costs = cls.calculate_costs(price, product_amount, total_volume, cargo)

            for i in range(len(bts)):
                cls.add_button_text(bts[i], costs[i])

    @classmethod
    def calculate_density(
        cls,
        product_amount: float,
        amount_in_box: float,
        box_weight: float,
        height: float,
        width: float,
        length: float,
    ):
        box_amount = ceil(product_amount / amount_in_box)
        origin_weight = box_amount * box_weight
        dimension = height * width * length
        volume = dimension * box_amount / 1000000
        total_weight = origin_weight + (volume / 2 * 50)
        total_volume = volume + (volume / 2 * 0.5)
        density = total_weight / total_volume

        return ceil(density), round(total_weight, 2), round(total_volume, 2)

    @classmethod
    def calculate_cargo(cls, density, total_weight, total_volume, currency):
        if density >= 100:
            cargo_dict, discount = cls.read_cargo_data()

            while str(density) not in cargo_dict.keys():
                density -= 1

            cargo_coefficient = float(cargo_dict[str(density)]) - float(discount)
            cargo = cargo_coefficient * total_weight * currency
        else:
            cargo = density * total_volume

        return ceil(cargo)

    @classmethod
    def calculate_costs(cls, price, product_amount, total_volume, cargo):
        """Calculete costs"""
        goods_cost = round(price * product_amount, 2)
        package = round(total_volume * 210, 2)
        insurance = round(goods_cost / 100, 2)
        total_cost = round(goods_cost + package + cargo + insurance, 2)
        one_goods_cost = round(total_cost / product_amount, 2)
        return goods_cost, package, insurance, total_cost, one_goods_cost, cargo

    @classmethod
    def read_cargo_data(cls):
        with open("data.json", "r") as file:
            json_to_dict = json.load(file)
            cargo_dictionary = json_to_dict["coefficient"]
            discount = json_to_dict["discount"]
            return cargo_dictionary, discount

    @staticmethod
    def add_button_text(bt, cost):
        text = f"{cost:.0f}" if cost.is_integer() else f"{round(cost, 2)}"
        bt.configure(text=f"{text} ¥")


class Frame(Calculation):
    # tuple[variable_name, text, row, column, button]
    labels: list[tuple[str, str, int, int, int]] = [
        ("price", "Цена в Китае", 0, 0, 0),
        ("product_amount", "Количество", 0, 1, 0),
        ("amount_in_box", "Кол. шт. в коробке", 0, 2, 0),
        ("height", "Высота", 2, 0, 0),
        ("width", "Ширина", 2, 1, 0),
        ("length", "Длина", 2, 2, 0),
        ("box_weight", "Вес", 4, 0, 0),
        ("currency", "Курс $/¥", 4, 1, 0),
        ("goods_cost", "Стоимость товара", 6, 0, 1),
        ("package", "Упаковка", 6, 1, 1),
        ("insurance", "Страховка", 6, 2, 1),
        ("total_cost", "Стоимость груза", 8, 0, 1),
        ("one_goods", "За единицу", 8, 1, 1),
        ("cargo", "Логистика", 8, 2, 1),
    ]
    entries: dict = {}
    buttons: dict = {}

    def __new__(cls):
        win = cttk.CTk()
        cttk.set_appearance_mode("light")
        cttk.set_default_color_theme("dark-blue")

        win.title("")
        win.geometry("660x650")
        win.grid_columnconfigure((0, 1, 2), weight=1)
        cls.create_fields(win, cls.labels)
        return win

    @classmethod
    def create_fields(cls, win, labels: list[tuple]) -> None:
        for name, text, row, column, button in labels:
            cls.create_label(win, text, row, column)
            row += 1
            if button:
                cls.create_button(win, name, row, column)
            else:
                cls.create_entry(win, name, row, column)

        cls.update_currency_entry()
        cls.create_calculate_button(win)

    @classmethod
    def create_label(cls, win, text, row, column):
        lb = cttk.CTkLabel(win, text=text, anchor="w")
        lb.grid(row=row, column=column, sticky="we", padx=(25, 0))
        return lb

    @classmethod
    def create_entry(cls, win, name: str, row: int, column: int, validate: bool = True):
        entry = cttk.CTkEntry(win)
        if validate:
            cls.entries[name] = entry
        entry.grid(row=row, column=column, sticky="we", padx=(25, 0))
        entry.bind("<FocusIn>", lambda event: cls.clean(entry, event))
        return entry

    @classmethod
    def create_button(cls, win, name: str, row: int, column: int) -> None:
        bt = cttk.CTkButton(
            win,
            text="─",
            fg_color="#e8e8e8",
            text_color="black",
            hover_color="#bababa",
        )
        cls.buttons[name] = bt
        bt.grid(row=row, column=column, padx=(25, 0), sticky="we")
        bt.configure(command=lambda: cls.copy_to_clip(bt.cget("text")))

    @classmethod
    def update_currency_entry(cls) -> None:
        currency = cls.get_currency()
        currency = currency if currency else "7.25"
        entry = cls.entries["currency"]
        entry.insert(0, currency)

    @classmethod
    def create_calculate_button(cls, win) -> None:
        button = cttk.CTkButton(
            win,
            text="Рассчитать",
            command=lambda: cls.calculate(win, cls.entries, cls.buttons),
            hover_color="#389147",
            text_color="white",
            fg_color="#3fa24f",
        )
        button.grid(row=5, column=2, padx=(25, 0), sticky="we")

    @classmethod
    def import_excel(cls):
        entries = cls.entries
        buttons = cls.buttons

        if cls.validate(win, entries):
            product_name = simpledialog.askstring("Товар", "Введите название товара:")
            if product_name:
                data = [product_name]
            else:
                return

            data = data + [entry.get() for entry in entries.values()]
            data = data + [button.cget("text")[:-2] for button in buttons.values()]

            try:
                file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
                if file:
                    workbook = openpyxl.load_workbook(file)
                    sheet = workbook.active
                    last_column = sheet.max_column + 1

                    for row_num, data_item in enumerate(data, start=1):
                        sheet.cell(row=row_num, column=last_column, value=data_item)
                    workbook.save(file)
                    workbook.close()
                    info = ("Успех", "Данные успешно импортированы в Excel!")
                    messagebox.showinfo(*info)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")

        else:
            messagebox.showerror("Ошибка", "Произведите рассчеты")

    @staticmethod
    def get_currency() -> Optional[str]:
        url = "https://api.freecurrencyapi.com/v1/latest"
        api_key = "fca_live_CKKpxzepp3RFgAQJD6EPR7TdYphkAwRkO9C6sLOQ"
        currencies = "CNY"

        params = {"apikey": api_key, "currencies": currencies}
        response = requests.get(url, params=params)

        if response.status_code == 200:
            json_data = response.json()["data"]
            return str(round(json_data["CNY"], 2))

    @staticmethod
    def copy_to_clip(text: str) -> None:
        if text and text != " ":  # ctk button behaivior
            copy(text[:-2])


class Table(Frame):
    columns: Final = ("No", "Density", "Coefficient")
    tree_position: Final = {
        "row": 10,
        "column": 0,
        "columnspan": 3,
        "sticky": "nsew",
        "padx": (20, 0),
        "pady": (20, 10),
    }

    def __new__(cls, win):
        cls.cargo_data, cls.discount_value = cls.read_cargo_data()

        cls.density = cttk.IntVar(value=None)
        cls.coefficient = cttk.DoubleVar(value=None)
        cls.discount = cttk.DoubleVar(value=None)

        cls.tree = cls.__create_tree_view(win)
        cls.__create_cargo_fields(win)
        cls.__create_buttons(win)

    @classmethod
    def __create_tree_view(cls, win):
        columns = cls.columns
        tree_position = cls.tree_position

        tree = ttk.Treeview(win, columns=columns, show="headings")
        tree.bind("<<TreeviewSelect>>", cls.select_item)
        tree.grid(tree_position)

        ttk.Style().theme_use("default")

        for column in columns:
            tree.heading(column, text=column, anchor=cttk.CENTER)
            tree.column(column, minwidth=25)

        tree.heading("No", text="No", anchor=cttk.CENTER)
        tree.heading("Density", text="Плотность", anchor=cttk.CENTER)
        tree.heading("Coefficient", text="Коэффициент", anchor=cttk.CENTER)

        scrollbar = cttk.CTkScrollbar(win, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=10, column=3, sticky="nsw", pady=(20, 10), padx=(0, 10))

        cls.load_table(tree)
        return tree

    @classmethod
    def load_table(cls, tree):
        cargo = cls.cargo_data
        for i, k in enumerate(cargo):
            tree.insert("", i, text=i + 1, values=(i + 1, k, cargo[k]))

    @classmethod
    def save_cargo(cls):
        tree = cls.tree
        if tree.selection():
            coefficient = cls.coefficient_entry.get()
            discount = cls.discount_entry.get()

            if "," in coefficient:
                coefficient = coefficient.replace(",", ".")
                cls.coefficient_entry.delete(0, cttk.END)
                cls.coefficient_entry.insert(0, coefficient)

            # Type of keys is string
            selected_density = str(cls.density.get())
            selected_coefficient = cls.cargo_data[selected_density]

            if selected_coefficient != coefficient:
                cls.cargo_data[selected_density] = coefficient

            with open("data.json", "w") as file:
                dictionary = {"coefficient": cls.cargo_data, "discount": discount}
                json.dump(dictionary, file)

            cls.tree.delete(*tree.get_children())
            cls.load_table(tree)

            cls.density.set(0)
            cls.coefficient.set(0)
            cls.discount.set(0)
            win.focus_set()

    @classmethod
    def __create_cargo_fields(cls, win):
        cls.create_label(win, "Плотность", 11, 0)
        lb = cls.create_label(win, "No", 12, 0)
        CENTER = cttk.CENTER
        lb.configure(fg_color="light gray", textvariable=cls.density, anchor=CENTER)

        cls.create_label(win, "Коэффициент", 11, 1)
        cls.coefficient_entry = cls.create_entry(win, "", 12, 1, validate=False)
        cls.coefficient_entry.configure(textvariable=cls.coefficient)

        cls.create_label(win, "Скидка", 11, 2)
        cls.discount_entry = cls.create_entry(win, "", 12, 2, validate=False)
        cls.discount_entry.configure(textvariable=cls.discount)

    @classmethod
    def __create_buttons(cls, win):
        save_bt = cttk.CTkButton(win, text="Сохранить", command=cls.save_cargo)
        save_bt.grid(row=15, column=0, padx=(25, 0), pady=15, sticky="we")

        text = "Импортировать в Excel"
        import_bt = cttk.CTkButton(win, text=text, command=cls.import_excel)
        import_bt.grid(row=15, column=1, padx=(25, 0), pady=15, sticky="we")

        cancel_bt = cttk.CTkButton(win, text="Отменить", command=cls.deselect_item)
        cancel_bt.grid(row=15, column=2, padx=(25, 0), pady=15, sticky="we")

    @classmethod
    def select_item(cls, event):
        if type(event) == Event:
            selected_item = cls.tree.selection()

            if selected_item:
                item = cls.tree.item(selected_item[0])
                values = item["values"]
                cls.density.set(int(values[1]))
                cls.coefficient.set(float(values[2]))
                cls.discount.set(float(cls.discount_value))

                # Focusing on the field with the coefficient
                cls.coefficient_entry.focus_set()
                cls.coefficient_entry.icursor(cttk.END)

    @classmethod
    def deselect_item(cls):
        selected_item = cls.tree.selection()
        if selected_item:
            cls.tree.selection_remove(selected_item)
            cls.density.set(0)
            cls.coefficient.set(0)
            cls.discount.set(0)
            win.focus_set()


if __name__ == "__main__":
    win = Frame()
    Table(win)
    win.mainloop()
